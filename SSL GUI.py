import tkinter as tk
from tkinter import ttk
import os
from PIL import ImageTk,Image
import customtkinter
from tkinter import filedialog
import openpyxl,xlrd
from openpyxl import Workbook
import numpy as np
import matplotlib.pyplot as plt
from scipy.io import wavfile
from scipy.signal import fftconvolve
import IPython
import pyroomacoustics as pra

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("SSL Results")
        self.geometry("1100x580")

        #set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((0,2), weight=0)

        self.tabview_1 = customtkinter.CTkTabview(self, width=200, height=70)
        self.tabview_1.grid(pady=10, padx=10)
        self.tabview_1.add("Steps")
        self.tabview_1.add("Example results")

        #load images with light and dark mode images
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "test_images")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "CustomTkinter_logo_single.png")), size=(26, 26))
        self.image_icon_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "image_icon_light.png")), size=(20, 20))
        self.folder_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "folder.jpg")), size=(20,20))
        self.python_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "python.jpg")), size=(20,20))
        self.upm_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "upm.jpg")), size=(250, 100))
        self.algo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "algorithm.png")), size=(20,20))
        self.save_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "save.png")), size=(20,20))
        self.sound_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "sound.png")), size=(20,20))
        self.music_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "music.png")), size=(20,20))
        self.cssm_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "cssm.png")), size=(20,20))
        self.tops_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "tops.png")), size=(20,20))
        self.frida_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "frida.png")), size=(20,20))
        self.waves_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "waves.png")), size=(20,20))

        #create navigation frame
        self.navigation_frame = customtkinter.CTkFrame(self, corner_radius=0)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")
        self.navigation_frame.grid_rowconfigure(5, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self.navigation_frame, text=" User Information", image=self.python_image,
                                                             compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.name_label = customtkinter.CTkLabel(self.navigation_frame, text="Please enter name: ", font=customtkinter.CTkFont(size=15))
        self.name_label.grid(row=1, column=0)
        self.name_entry = customtkinter.CTkEntry(self.navigation_frame, width=50, height=10, border_width=2)
        self.name_entry.grid(row=2, column=0)

        self.course_label = customtkinter.CTkLabel(self.navigation_frame, text="Please enter course: ", font=customtkinter.CTkFont(size=15))
        self.course_label.grid(row=3, column=0)
        self.course_entry = customtkinter.CTkEntry(self.navigation_frame, width=100, height=10, border_width=2)
        self.course_entry.grid(row=4, column=0)

        self.save_button = customtkinter.CTkButton(self.navigation_frame, height=40, width=60, text="Save ",
                                                   image=self.save_image, compound="left", command=self.save)
        self.save_button.grid(row=6, column=0)

        self.appearance_mode_menu = customtkinter.CTkOptionMenu(self.navigation_frame, values=["Light", "Dark", "System"],
                                                                command=self.change_appearance_mode_event)
        self.appearance_mode_menu.grid(row=7, column=0, padx=20, pady=20, sticky="s")

        self.upm_image_label = customtkinter.CTkLabel(self.navigation_frame, text="", image=self.upm_image, compound="center")
        self.upm_image_label.grid(row=8, column=0, sticky="ns")


        # create home frame        
        home_frame = tk.Frame(self, bg='#9797ff')

        lbl_sound_path = tk.Label(home_frame, text='Sound Path:', padx=20, pady=10, font=('verdana',15), bg='#9797ff')
        lbl_show_sound = tk.Label(home_frame, bg='#9797ff')
        lbl_show_sound_ori = tk.Label(home_frame, bg='#9797ff')
        lbl_show_sound_fil = tk.Label(home_frame, bg='#9797ff')
        option_sound_list = ["1kHz Sine Waves", "Pink Noise", "Recorded Speech", "Hand Claps", "Birds Chirping"]
        value_inside_option_sound = tk.StringVar(home_frame)
        value_inside_option_sound.set("Select an option")
        option_sound = tk.OptionMenu(home_frame, value_inside_option_sound, *option_sound_list)
        lbl_position = tk.Label(home_frame, text='Position:', padx=20, pady=10, font=('verdana',15), bg='#9797ff')
        option_position_list = ["1", "2", "3"]
        value_inside_option_position = tk.StringVar(home_frame)
        value_inside_option_position.set("Select an option")
        option_position = tk.OptionMenu(home_frame, value_inside_option_position, *option_position_list)
        btn_browse = tk.Button(home_frame, text='Select Sound',bg='grey', fg='#ffffff', font=('verdana',15))
        music_button = tk.Button(home_frame, text="MUSIC", bg='blue',fg='#ffffff', font=('verdana',15))

        def selectPic():
            global img
            global img_ori
            global img_fil
            file = value_inside_option_sound.get()
            position = value_inside_option_position.get()
            if file=="1kHz Sine Waves":
                picture = "C:/Pictures/Pure tones.png"
                img = Image.open(picture)
                img = img.resize((700,150), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                lbl_show_sound['image'] = img
                disc = tk.Label(home_frame, text='This is the waveform and spectogram of 1kHz Sine Waves',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                disc.grid(row=6, column=0)
                if position=="1":
                    #ori
                    picture_ori = "C:/Pictures/1/ori/1kHz Sine Waves.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with 1kHz Sine Waves',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/1/filtered/1kHz Sine Waves.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source', pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                elif position=="2":
                    #ori
                    picture_ori = "C:/Pictures/2/ori/1kHz Sine Waves.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with 1kHz Sine Waves',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/2/filtered/1kHz Sine Waves.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                else:
                    #ori
                    picture_ori = "C:/Pictures/3/ori/1kHz Sine Waves.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with 1kHz Sine Waves',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/3/filtered/1kHz Sine Waves.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
            elif file=="Pink Noise":
                picture = "C:/Pictures/Broadband noise.png"
                img = Image.open(picture)
                img = img.resize((700,150), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                lbl_show_sound['image'] = img
                disc = tk.Label(home_frame, text='This is the waveform and spectogram of Pink Noise',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                disc.grid(row=6, column=0)
                if position=="1":
                    #ori
                    picture_ori = "C:/Pictures/1/ori/Pink Noise.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Pink Noise',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/1/filtered/Pink Noise.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                elif position=="2":
                    #ori
                    picture_ori = "C:/Pictures/2/ori/Pink Noise.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Pink Noise',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/2/filtered/Pink Noise.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                else:
                    #ori
                    picture_ori = "C:/Pictures/3/ori/Pink Noise.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Pink Noise',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/3/filtered/Pink Noise.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
            elif file=="Recorded Speech":
                picture = "C:/Pictures/Speech signals.png"
                img = Image.open(picture)
                img = img.resize((700,150), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                lbl_show_sound['image'] = img
                disc = tk.Label(home_frame, text='This is the waveform and spectogram of Recorded Speech',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                disc.grid(row=6, column=0)
                if position=="1":
                    #ori
                    picture_ori = "C:/Pictures/1/ori/Recorded Speech.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Recorded Speech',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/1/filtered/Recorded Speech.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                elif position=="2":
                    #ori
                    picture_ori = "C:/Pictures/2/ori/Recorded Speech.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Recorded Speech',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/2/filtered/Recorded Speech.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                else:
                    #ori
                    picture_ori = "C:/Pictures/3/ori/Recorded Speech.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Recorded Speech',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/3/filtered/Recorded Speech.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
            elif file=="Hand Claps":
                picture = "C:/Pictures/Impulsive sounds.png"
                img = Image.open(picture)
                img = img.resize((700,150), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                lbl_show_sound['image'] = img
                disc = tk.Label(home_frame, text='This is the waveform and spectogram of Hand Claps',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                disc.grid(row=6, column=0)
                if position=="1":
                    #ori
                    picture_ori = "C:/Pictures/1/ori/Hand Claps.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Hand Claps',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/1/filtered/Hand Claps.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                elif position=="2":
                    #ori
                    picture_ori = "C:/Pictures/2/ori/Hand Claps.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Hand Claps',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/2/filtered/Hand Claps.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                else:
                    #ori
                    picture_ori = "C:/Pictures/3/ori/Hand Claps.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Hand Claps',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/3/filtered/Hand Claps.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
            else:
                picture = "C:/Pictures/Environmental sounds.png"
                img = Image.open(picture)
                img = img.resize((700,150), Image.ANTIALIAS)
                img = ImageTk.PhotoImage(img)
                lbl_show_sound['image'] = img
                disc = tk.Label(home_frame, text='This is the waveform and spectogram of Birds Chirping',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                disc.grid(row=6, column=0)
                if position=="1":
                    #ori
                    picture_ori = "C:/Pictures/1/ori/Birds Chirping.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Birds Chirping',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/1/filtered/Birds Chirping.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                elif position=="2":
                    #ori
                    picture_ori = "C:/Pictures/2/ori/Birds Chirping.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Birds Chirping',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/2/filtered/Birds Chirping.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)
                else:
                    #ori
                    picture_ori = "C:/Pictures/3/ori/Birds Chirping.png"
                    img_ori = Image.open(picture_ori)
                    img_ori = img_ori.resize((700,150), Image.ANTIALIAS)
                    img_ori = ImageTk.PhotoImage(img_ori)
                    lbl_show_sound_ori['image'] = img_ori
                    disc = tk.Label(home_frame, text='Sound source from drone with Birds Chirping',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=8, column=0)
                    #fil
                    picture_fil = "C:/Pictures/3/filtered/Birds Chirping.png"
                    img_fil = Image.open(picture_fil)
                    img_fil = img_fil.resize((700,150), Image.ANTIALIAS)
                    img_fil = ImageTk.PhotoImage(img_fil)
                    lbl_show_sound_fil['image'] = img_fil
                    disc = tk.Label(home_frame, text='Filtered sound source',padx=20, pady=10, font=('verdana', 10), bg='#9797ff')
                    disc.grid(row=10, column=0)

        def execute():
            file = value_inside_option_sound.get()
            position = value_inside_option_position.get()
            if file=="1kHz Sine Waves":
                if position=="1":
                    opti = tk.Label(self.function_frame, text='-1.9714',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='-1.2751',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='35.32',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                elif position=="2":
                    opti = tk.Label(self.function_frame, text='0.2397',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='0.3789',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='58.10',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                else:
                    opti = tk.Label(self.function_frame, text='1.6022',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='2.0178',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='25.94',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
            elif file=="Pink Noise":
                if position=="1":
                    opti = tk.Label(self.function_frame, text='-1.9714',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='-1.6915',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='14.20',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                elif position=="2":
                    opti = tk.Label(self.function_frame, text='0.2397',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='0.1604',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='33.07',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                else:
                    opti = tk.Label(self.function_frame, text='1.6022',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='1.8225',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='13.75',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
            elif file=="Recorded Speech":
                if position=="1":
                    opti = tk.Label(self.function_frame, text='-1.9714',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='-1.7852',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='9.44',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                elif position=="2":
                    opti = tk.Label(self.function_frame, text='0.2397',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='0.2744',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='14.50',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                else:
                    opti = tk.Label(self.function_frame, text='1.6022',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='1.7063',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='6.50',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
            elif file=="Hand Claps":
                if position=="1":
                    opti = tk.Label(self.function_frame, text='-1.9714',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='-1.3657',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='30.72',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                elif position=="2":
                    opti = tk.Label(self.function_frame, text='0.2397',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='0.3663',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='52.84',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                else:
                    opti = tk.Label(self.function_frame, text='1.6022',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='1.9003',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='18.61',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
            else:
                if position=="1":
                    opti = tk.Label(self.function_frame, text='-1.9714',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='-2.2668',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='14.99',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                elif position=="2":
                    opti = tk.Label(self.function_frame, text='0.2397',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='0.3079',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='28.47',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)
                else:
                    opti = tk.Label(self.function_frame, text='1.6022',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=7, column=1)
                    opti = tk.Label(self.function_frame, text='1.8945',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=8, column=1)
                    opti = tk.Label(self.function_frame, text='18.24',padx=20, pady=10, font=('verdana', 10))
                    opti.grid(row=9, column=1)

        music_button['command'] = execute

        btn_browse['command'] = selectPic

        home_frame.grid(row=0, column=1, sticky = "nsew")
        home_frame.grid_columnconfigure(0, weight=1)

        lbl_sound_path.grid(row=0, column=0)
        option_sound.grid(row=1, column=0)
        lbl_show_sound.grid(row=5, column=0)
        lbl_show_sound_ori.grid(row=7, column=0)
        lbl_show_sound_fil.grid(row=9, column=0)
        btn_browse.grid(row=4, column=0, padx=10, pady=10)
        lbl_position.grid(row=2, column=0)
        option_position.grid(row=3, column=0)
        music_button.grid(row=11, column=0)

        #create function frame
        self.function_frame = customtkinter.CTkFrame(self, corner_radius=0)

        self.function_frame_label = customtkinter.CTkLabel(self.function_frame, text=" Algorithm Selection", image=self.algo_image,
                                                           compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        #MUSIC
        self.music_button = customtkinter.CTkButton(self.function_frame, height=40, width=60, text="MUSIC",
                                                   image=self.music_image, compound="left")
        
        #CSSM
        self.cssm_button = customtkinter.CTkButton(self.function_frame, height=40, width=60, text="CSSM",
                                                   image=self.cssm_image, compound="left")
        
        #WAVES
        self.waves_button = customtkinter.CTkButton(self.function_frame, height=40, width=60, text="WAVES",
                                                   image=self.waves_image, compound="left")
        
        #TOPS
        self.tops_button = customtkinter.CTkButton(self.function_frame, height=40, width=60, text="TOPS",
                                                   image=self.tops_image, compound="left")
        
        #FRIDA
        self.frida_button = customtkinter.CTkButton(self.function_frame, height=40, width=60, text="FRIDA",
                                                   image=self.frida_image, compound="left")
        
        self.opti_label = customtkinter.CTkLabel(self.function_frame, text="OptiTrack's data (x-axis): ", font=customtkinter.CTkFont(size=15))

        self.cal_label = customtkinter.CTkLabel(self.function_frame, text="Calculated Azimuth (x-axis): ", font=customtkinter.CTkFont(size=15))

        self.percent_label = customtkinter.CTkLabel(self.function_frame, text="Percentage Error (%): ", font=customtkinter.CTkFont(size=15))

        self.function_frame.grid(row=0, column=2, sticky="nsew")
        self.function_frame.grid_rowconfigure(6, weight=1)
        self.function_frame.grid_rowconfigure(10, weight=1)

        self.function_frame_label.grid(row=0, column=0, padx=20, pady=20)

        self.music_button.grid(row=1, column=0)
        self.cssm_button.grid(row=2, column=0)
        self.waves_button.grid(row=3, column=0)
        self.tops_button.grid(row=4, column=0)
        self.frida_button.grid(row=5, column=0)

        self.opti_label.grid(row=7, column=0)
        self.cal_label.grid(row=8, column=0)
        self.percent_label.grid(row=9, column=0)
                
    def save(self):
        print("{}".format(self.name_entry.get()))
        print("{}".format(self.course_entry.get()))

    def change_appearance_mode_event(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def pyroom(self):
        ### Direction of Arrival

        ############### Location of sources ###########

        dire = [90] #,90,135,180,225,270,315,360]

        for i in dire:
            #deg = float(input("Enter the Direction in Degrees ..'))
            distance = 1.4 # meters
            azimuth = (int(i)) / 180. * np.pi # Put in the Location of the Source (X,Z) # this shows two damn azimuths
            print('True Azimuth in rad is ..',azimuth)
            azdeg = (azimuth/np.pi )*180
            print('True Azimuth in deg is ..',azdeg)
            print()
            #distance = 1.8096778885935059 # meters

            # Constants
            c = 343. # speed of sound
            axis = 1
            #fs=32000 # sampling frequency
            nfft=256 # FFT size
            freq_range=[300, 3500]

            fs, signal2 =wavfile.read("C:/Users/User/Documents/Engineer UPM/Sem 7/Final Year Project/Results/(filtered) 1kHz Sine Waves.wav") # Use this 'Drone and Common Nighthawk (Combined).wav'

            snr_db = 5 # signal-to-noise ratio
            sigma2 = 10**(-snr_db / 10)/(4. * np.pi * distance)**2

            # Create an anechoic room
            room_dim= np.r_[6.11,9.4742] # in meters
            aroom = pra.ShoeBox(room_dim, fs=fs, max_order=0, sigma2_awgn=sigma2)
            sigma2_awgn=sigma2

            # Array geometry
            echo = pra.circular_2D_array(center=room_dim/2, M=4, phi0=0, radius=30e-3)
            #echo = np.concatenate((echo, np.array(room_dim/2, ndmin=2).T), axis=1)
            aroom.add_microphone_array(pra.MicrophoneArray(echo, aroom.fs) )
            #fig, ax = aroom.plot

            # Add sources of 1 second duration
            #rng = np.random.RandomState(23)# GAuss
            # duration_samples = int(fs)

            ang = azimuth
            source_location = room_dim/2+ distance* np.r_[np.cos(ang),np.sin(ang)]
            source_signal = signal2
            aroom.add_source(source_location, signal = source_signal)

            # Run the simulation
            aroom.simulate()
            
            # DOA algorithms require an STFT input
            X = pra.transform.stft.analysis(aroom.mic_array.signals.T,nfft,nfft // 2)
            X = X.transpose([2, 1, 0])

            # Comparing Alogrithms
            algo_name = 'MUSIC'
            spatial_resp = dict()

            # Construct the new DOA object

            doa = pra.doa.algorithms [algo_name](echo, fs, nfft, c=c, num_src=1)

            # this call here perform localization on the frames in X
            doa.locate_sources(X, freq_range=freq_range)

            # store spatial response
            spatial_resp[algo_name] = doa.grid.values
            print(doa.dict)
            ()
            # normalize
            min_val = spatial_resp[algo_name].min()
            max_val = spatial_resp[algo_name].max()
            spatial_resp[algo_name]=(spatial_resp[algo_name] - min_val)/(min_val - max_val)

            # Plot the estimated spatial spectra and compare it with the true locations!
            # plotting param base=0
            height = 10
            true_col = [0,0,0] # The Colouring of Lines

            # plot
            phi_plt = doa.grid.azimuth
            fig = plt.figure()
            ax= fig.add_subplot(111, projection='polar')
            c_phi_plt = np.r_[phi_plt]

            c_dirty_img = np.r_[spatial_resp[algo_name]]
            ax.plot(c_phi_plt, base+height * c_dirty_img, linewidth=3,alpha=0.55, linestyle='-', label= 'spatial spectrum')
            plt.title(algo_name)

            # plot true col

            ax.plot([ang], [base + height], linewidth=3, linestyle='--',color=true_col, alpha=0.6)

            K=1
            ax.scatter(azimuth, base + height* np.ones(K), c=np.tile(true_col,(K, 1)), s=500, alpha=0.75, marker='*', linewidths=0,label='true locations')

            plt.legend()
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels, framealpha=0.5,scatterpoints=1, loc='center right', fontsize=16, ncol = 1, bbox_to_anchor=(1.6, 0.5), handletextpad=.2, columnspacing=1.7, labelspacing=0.1)

            ax.set_xticks(np.linspace(0, 2* np.pi, num=12, endpoint=False))
            ax.xaxis.set_label_coords(0.5, -0.11)
            ax.set_yticks(np.linspace(0, 1, 2))
            ax.xaxis.grid(visible=True, color=[0.3, 0.3, 0.3], linestyle=':')
            ax.yaxis.grid(visible=True, color=[0.3, 0.3, 0.3], linestyle='--')
            ax.set_ylim ([0, 1.05 * (base + height)])

            plt.show()
            print()
        
if __name__ == "__main__":
    app = App()
    app.mainloop()

##        self.home_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
##        self.home_frame.grid_columnconfigure(0, weight=1)
##
##        self.image_path = customtkinter.CTkLabel(self.home_frame, text="Image Path: ", font=customtkinter.CTkFont(size=15, weight="bold"))
##        self.show_pic_label = customtkinter.CTkLabel(self.home_frame, fg_color="transparent")
##        self.entry_pic_path = customtkinter.CTkEntry(self.home_frame, font=ustomtkinter.CTkFont(size=15))
##        self.folder_button = customtkinter.CTkButton(self.home_frame, height=40, border_spacing=10, text="Select folder",
##                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
##                                                   image=self.folder_image, anchor="w")
##
##        def open(self):
##            global img
##            self.filename = filedialog.askopenfilename(initialdir="/Results", title="Select a file", filetypes=(("png files","*.wav"),("all files", "*.*")))
##            self.img = customtkinter.CTkImage(Image.open(self.filename))
##            self.img = self.img.resize((200,200), cTkImage.ANTIALIAS)
##            self.img = 
##            self.img = customtkinter.CTkLabel(master=self.home_frame, text="", image=self.music).pack()
##            self.music.grid(row=1, column=0)
##            filename = filedialog.askopenfilename(initialdir="/Results", title="Select a file", filetypes=(("png files","*.wav"),("all files", "*.*")))
##            music = customtkinter.CTkImage(Image.open(filename))
##            music = music.resize((100,100), Image.ANTIALIAS)
##            music = customtkinter.CTkImage.PhotoImage(music)
##        
##        self.music_label['music']=music
##        self.music_label = customtkinter.CTkLabel(self.home_frame, text="", image=music)
##        self.music_label.grid(row=1, column=0)
##
##        self.folder_button['command']=open
##
##        self.home_frame_label = customtkinter.CTkLabel(self.home_frame,compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
##        self.home_frame_label.grid(row=0, column=0, padx=20, pady=20)
##
##        self.folder_button = customtkinter.CTkButton(master=self.home_frame, corner_radius=0, height=40, border_spacing=10, text="Select folder",
##                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
##                                                   image=self.folder_image, anchor="w", command=self.open)
##        self.folder_button.grid(row=1, column=0)
##
##        self.music_label = customtkinter.CTkLabel(self.home_frame, text="", image=self.pure_image)
##        self.music_label.grid(row=1, column=0, sticky="ns")

##        self.opti_entry.configure(state=tk.DISABLED)
##        self.cal_entry.configure(state=tk.DISABLED)
##        self.percent_entry.configure(state=tk.DISABLED)
##
##        excel_path = "C:/Users/User/Documents/Engineer UPM/Sem 7/Final Year Project/OptiTrack dataset.xlsx"
##        def execute():
##            file = value_inside_option_sound.get()
##            position = value_inside_option_position.get()
##            self.opti_entry.configure(state=tk.NORMAL)
##            self.cal_entry.configure(state=tk.NORMAL)
##            self.percent_entry.configure(state=tk.NORMAL)
##
##            self.opti_entry.delete(0, 'end')
##            self.cal_entry.delete(0, 'end')
##            self.percent_entry.configure(0, 'end')
##
##            excel = openpyxl.load_workbook(excel_path)
##            sheet = excel['Sheet2']
##
##            for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=10, values_only=True):
##                if cell[0]==str(file):
##                    if position == "1":
##                        self.opti_entry.insert(0, cell[1])
##                        self.cal_entry.insert(0, cell[2])
##                        self.percent_entry.insert(0, cell[3])
##
##                        self.opti_entry.configure(state=tk.DISABLED)
##                        self.cal_entry.configure(state=tk.DISABLED)
##                        self.percent_entry.configure(state=tk.DISABLED)
##                    elif position == "2":
##                        self.opti_entry.insert(0, cell[4])
##                        self.cal_entry.insert(0, cell[5])
##                        self.percent_entry.insert(0, cell[6])
##
##                        self.opti_entry.configure(state=tk.DISABLED)
##                        self.cal_entry.configure(state=tk.DISABLED)
##                        self.percent_entry.configure(state=tk.DISABLED)
##                    else:
##                        self.opti_entry.insert(0, cell[7])
##                        self.cal_entry.insert(0, cell[8])
##                        self.percent_entry.insert(0, cell[9])
##
##                        self.opti_entry.configure(state=tk.DISABLED)
##                        self.cal_entry.configure(state=tk.DISABLED)
##                        self.percent_entry.configure(state=tk.DISABLED)

